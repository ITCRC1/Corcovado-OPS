"""
Genera reportes descargables (Excel y PDF) a partir de datos tabulares.
Usado por los endpoints /api/export/* en main.py.
"""
import io
import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

FOREST = "2C4A3E"
FOREST_RGB = colors.HexColor("#2C4A3E")
SAND_RGB = colors.HexColor("#F5F1E8")

_resource_dir = os.environ.get("HOTEL_RESOURCE_DIR") or os.path.join(os.path.dirname(__file__), "..", "frontend")
LOGO_PATH = os.path.join(_resource_dir if os.environ.get("HOTEL_RESOURCE_DIR") else os.path.join(os.path.dirname(__file__), ".."), "frontend", "assets", "logo.jpg")
if not os.path.exists(LOGO_PATH):
    LOGO_PATH = os.path.join(os.path.dirname(__file__), "..", "frontend", "assets", "logo.jpg")


def to_xlsx(columns, rows, title):
    """columns: lista de (clave, encabezado). rows: lista de dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Reporte"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True, color=FOREST)
    ws.cell(row=2, column=1, value=f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, italic=True, color="6E6A5C")

    header_row = 4
    for i, (key, label) in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=i, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=FOREST)
        c.alignment = Alignment(horizontal="left")

    for r, row in enumerate(rows, start=header_row + 1):
        for i, (key, label) in enumerate(columns, start=1):
            val = row.get(key, "")
            if val is None:
                val = ""
            ws.cell(row=r, column=i, value=val)
            if r % 2 == 0:
                ws.cell(row=r, column=i).fill = PatternFill("solid", fgColor="F5F1E8")

    for i, (key, label) in enumerate(columns, start=1):
        max_len = max([len(str(label))] + [len(str(row.get(key, ""))) for row in rows]) if rows else len(str(label))
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def to_pdf(columns, rows, title, subtitle=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(letter),
        topMargin=0.6 * inch, bottomMargin=0.6 * inch, leftMargin=0.5 * inch, rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleForest", parent=styles["Title"], textColor=FOREST_RGB, fontSize=16, spaceAfter=2)
    sub_style = ParagraphStyle("SubGray", parent=styles["Normal"], textColor=colors.HexColor("#6E6A5C"), fontSize=9)

    elements = []
    if os.path.exists(LOGO_PATH):
        try:
            elements.append(Image(LOGO_PATH, width=1.6 * inch, height=0.5 * inch))
            elements.append(Spacer(1, 8))
        except Exception:
            pass
    elements.append(Paragraph(title, title_style))
    if subtitle:
        elements.append(Paragraph(subtitle, sub_style))
    elements.append(Paragraph(f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", sub_style))
    elements.append(Spacer(1, 14))

    headers = [label for _, label in columns]
    data = [headers]
    for row in rows:
        data.append([str(row.get(key, "") if row.get(key, "") is not None else "") for key, _ in columns])

    if len(data) == 1:
        data.append(["Sin datos"] + [""] * (len(columns) - 1))

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), FOREST_RGB),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, SAND_RGB]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E3DFD3")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf
